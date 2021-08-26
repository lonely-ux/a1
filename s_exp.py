import openpyxl
import os, sys, subprocess
import pandas as pd
import tkinter as tk
import shutil
import urllib
import yfinance as yf
import statistics
from multiprocessing import Process
from pathlib import Path
from smb.SMBHandler import SMBHandler


def main():
    screen = tk.Tk()
    screen.geometry("720x500")
    screen.title("Finances")
    filename = Path("/path/to/file/Expenses.xlsx")
    finances_txt = Path("/path/to/file/finances.txt")

    #Cool thing i found on stack overflow where it "checks" what os you are on, and uses the correct method to launch the file correctly.
    def file_path():
        if sys.platform == "win32":
            os.startfile(filename)
        else:
            opener = "open" if sys.platform =="darwin" else "xdg-open"
            subprocess.call([opener, filename])
    
    def finances():
        if sys.platform == "win32":
            os.startfile(finances_txt)
        else:
            opener = "open" if sys.platform =="darwin" else "xdg-open"
            subprocess.call([opener, finances_txt])
    #Used this program with a home server to save and upload the xlsx and txt file to it.
    def upload_xlsx():
        file = open(filename, "rb")
        url = 'smb://username:password@192.168.1.18/path/to/file'
        smb = urllib.request.build_opener(SMBHandler)
        data = smb.open(url, data = file)
        data.close()


    def download_xlsx():
        url = 'smb://username:password@192.168.1.18/path/to/file'
        smb = urllib.request.build_opener(SMBHandler)
        fh = smb.open(url)
        with fh as response, open ("Expenses.xlsx", "wb") as out_file:
            shutil.copyfileobj(response, out_file)

    #Class for the calculations done
    class totals:
        def __init__(self):
            self.wb = openpyxl.load_workbook(filename)
            self.sheets = self.wb.sheetnames
            self.sheets_pd = pd.read_excel(filename, sheet_name=None)
            self.df = pd.read_excel(filename, self.sheets[-1])
            self.df_all = pd.concat(self.sheets_pd[frame] for frame in self.sheets_pd.keys())
            self.page = self.wb.active
            self.bill= round(self.df['Bills'].sum(), 2)
            self.total = round(self.df['Price'].sum(), 2)
            self.income = round(self.df['INCOME'].sum(), 2)
            self.tfsa_tot = round(self.df_all['TFSA'].sum(), 2)
            self.tfsa = round(self.df['TFSA'].sum(), 2)
            self.PERSONAL = round(self.df['PERSONAL'].sum(), 2)
            #Category Totals
            self.gas_tot = round(self.df.query('GAS==1')['Price'].sum(), 2)
            self.food_tot = round(self.df.query('FOOD==1')['Price'].sum(), 2)
            self.paybacks_tot = round(self.df.query('PAYBACKS==1')['Price'].sum(), 2)
            self.supplies_tot = round(self.df.query('SUPPLIES==1')['Price'].sum(), 2)
            self.luxury_tot = round(self.df.query('LUXURY==1')['Price'].sum(), 2)
            self.credit_tot = round(self.df.query('CREDIT==1')['Price'].sum(), 2)
            #Overall Totals
            self.bill_tot = self.bill - self.paybacks_tot
            self.necessity_tot = self.gas_tot + self.supplies_tot
            self.discret_tot = self.food_tot + self.luxury_tot
            self.investments = self.tfsa + self.PERSONAL
            self.grand_total = self.necessity_tot + self.discret_tot
            self.income_used = round(self.bill_tot / self.income * 100, 2)
            self.income_saved = (self.income_used - 100) * -1

        #This sets the result on the main GUI
        def main_total(self):
            var.e_7.set(str(self.bill_tot) + '$')
            var.e_8.set(str(self.necessity_tot) + '$')
            var.e_9.set(str(self.discret_tot) + '$')
            var.e_10.set(str(self.income_saved) + '%')
            var.e_11.set(str(self.income) + '$')
        #This sets the result on the basic overview GUI
        def overview(self):
            var.overview1.set(str(self.gas_tot) + '$')
            var.overview2.set(str(self.food_tot) + '$')
            var.overview3.set(str(self.paybacks_tot) + '$')
            var.overview4.set(str(self.supplies_tot) + '$')
            var.overview5.set(str(self.luxury_tot) + '$')
            var.overview6.set(str(self.credit_tot) + '$')
            var.overview7.set(str(self.income_saved) + '%')
            var.overview8.set(str(self.grand_total) + '$')
        ##This sets the result on the investments GUI
        def inv_total(self):
            var.inv3.set(str(self.tfsa + self.PERSONAL) + '$')
            var.inv4.set(str(self.tfsa_tot) + '$')


    t = totals()

    #lots of variables so i just tossed em all in here
    class variables:
        def __init__(self):
            #Overview Window
            self.overview1 = tk.DoubleVar(screen)
            self.overview2 = tk.DoubleVar(screen)
            self.overview3 = tk.DoubleVar(screen)
            self.overview4 = tk.DoubleVar(screen)
            self.overview5 = tk.DoubleVar(screen)
            self.overview6 = tk.DoubleVar(screen)
            self.overview7 = tk.DoubleVar(screen)
            self.overview8 = tk.DoubleVar(screen)
            #Main_Window
            self.e_1 = tk.StringVar(screen)
            self.e_2 = tk.StringVar(screen)
            self.e_3 = tk.DoubleVar(screen)
            self.e_4 = tk.StringVar(screen)
            self.e_5 = tk.StringVar(screen)
            self.e_6 = tk.DoubleVar(screen)
            self.e_7 = tk.StringVar(screen)
            self.e_8 = tk.StringVar(screen)
            self.e_9 = tk.StringVar(screen)
            self.e_10 = tk.StringVar(screen)
            self.e_11 = tk.StringVar(screen)
            self.e_13 = tk.DoubleVar(screen)
            self.check1 = tk.DoubleVar(screen)
            self.check2 = tk.DoubleVar(screen)
            self.check3 = tk.DoubleVar(screen)
            self.check4 = tk.DoubleVar(screen)
            self.check5 = tk.DoubleVar(screen)
            self.check6 = tk.DoubleVar(screen)
            #Investment Window
            self.inv1 = tk.DoubleVar(screen)
            self.inv2 = tk.DoubleVar(screen)
            self.inv3 = tk.DoubleVar(screen)
            self.inv4 = tk.DoubleVar(screen)
            self.inv5 = tk.StringVar(screen)
            self.inv5a = tk.DoubleVar(screen)
            self.inv6 = tk.DoubleVar(screen)


    var = variables()

    #class for all the minor GUIs
    class windows:
        def overview_window(self):
            window = tk.Toplevel(screen)
            window.geometry('500x400')
            window.title('Overview')
            L = tk.Label(window, text='Overview', font=('arial', 30, 'bold')).pack(side='top')
            L1 = tk.Label(window, text='GAS:', font=('arial', 15)).place(x=84, y=80)
            L2 = tk.Label(window, text='FOOD:', font=('arial', 15)).place(x=67, y=130)
            L3 = tk.Label(window, text='PAYBACKS:', font=('arial', 15)).place(x=23, y=180)
            L4 = tk.Label(window, text='SUPPLIES:', font=('arial', 15)).place(x=30, y=230)
            L5 = tk.Label(window, text='LUXURY:', font=('arial', 15)).place(x=47, y=280)
            L6 = tk.Label(window, text='CREDIT:', font=('arial', 15)).place(x=52, y=330)
            L7 = tk.Label(window, text=r'%_OF_INCOME_SAVED:', font=('arial', 12)).place(x=303, y=180)
            L8 = tk.Label(window, text='TOTAL_SPENT', font=('arial', 15)).place(x=328, y=90)
            E1 = tk.Entry(window, textvar=var.overview1, bd=5)
            E1.place(x=150, y=80)
            E2 = tk.Entry(window, textvar=var.overview2, bd=5)
            E2.place(x=150, y=130)
            E3 = tk.Entry(window, textvar=var.overview3, bd=5)
            E3.place(x=150, y=180)
            E4 = tk.Entry(window, textvar=var.overview4, bd=5)
            E4.place(x=150, y=230)
            E5 = tk.Entry(window, textvar=var.overview5, bd=5)
            E5.place(x=150, y=280)
            E6 = tk.Entry(window, textvar=var.overview6, bd=5)
            E6.place(x=150, y=330)
            E7 = tk.Entry(window, textvar=var.overview7, bd=5)
            E7.place(x=335, y=210)
            E8 = tk.Entry(window, textvar=var.overview8, bd=5)
            E8.place(x=335, y=120)
            B1 = tk.Button(window, text='SHOW_TOTAL', font=('arial', 15, 'bold'), command=t.overview).place(x=320, y=270)
            window.mainloop

        def investment_window(self):
            window = tk.Toplevel(screen)
            window.geometry('1100x780')
            window.title('Investments')
            L = tk.Label(window, text='Investments', font=('arial', 30, 'bold')).pack(side='top')
            L1 = tk.Label(window, text='TFSA', font=('arial', 15)).place(x=75, y=70)
            L2 = tk.Label(window, text='Personal', font=('arial', 15)).place(x=352, y=70)
            L3 = tk.Label(window, text='Invested_This_Month:', font=('arial', 15)).place(x=30, y=180)
            L4 = tk.Label(window, text='Total TFSA Contributions:', font=('arial', 15)).place(x=30, y=260)
            L5 = tk.Label(window, text='Ticker for Stock/Crypto:', font=('arial', 15)).place(x=30, y=570)
            L5a = tk.Label(window, text='Amount of Shares/Units:', font=('arial', 15)).place(x=30, y=620)
            E1 = tk.Entry(window, textvar=var.inv1, bd=5)
            E1.place(x=30, y=100)
            E2 = tk.Entry(window, textvar=var.inv2, bd=5)
            E2.place(x=320, y=100)
            E3 = tk.Entry(window, textvar=var.inv3, bd=5)
            E3.place(x=320, y=180)
            E4 = tk.Entry(window, textvar=var.inv4, bd=5)
            E4.place(x=320, y=260)
            T1 = tk.Text(window, bd=5)
            T1.place(x= 530, y=50, height = 700, width = 500)
            E5 = tk.Entry(window, textvar=var.inv5, bd=5)
            E5.place(x=320, y=570)
            E5a = tk.Entry(window, textvar=var.inv5a, bd=5)
            E5a.place(x=320, y=620)
            E6 = tk.Entry(window, textvar=var.inv6, bd=5)
            E6.place(x=200, y=670)

            def txt(): #this displays the txt file on the text widget in tkinter
                with open(finances_txt, 'r') as f:
                    T1.insert(tk.INSERT, f.read())

            def exit(): #Exit button, saves the file and closes window. This was more or less copypasted so I dont really understand it
                def save():
                    with open(finances_txt, "w") as f:
                        data = T1.get("1.0", tk.END)
                        f.write(data)
                ux = Process(target=save, name='Server txt')
                try:
                    ux.start()
                    ux.join(timeout=1)
                    ux.terminate()
                    window.destroy()
                except:
                    window.destroy()

            def enter(self): #Trys to upload the txt file to server
                info.server_c()

            def s_info(): #Searches very basic info on yahoo finance. case and symbol sensitive. Ex. ETH_CAD, AAPL. All entrys done in the console/terminal 
                ticker = input("stock, etf or crypto?: ")
                if ticker == "stock":
                    stock = input("Enter ticker: ")
                    data = yf.Ticker(stock)
                    c = data.info["profitMargins"]
                    b = data.info["sector"]
                    a = data.info["regularMarketPrice"]
                    y = data.info["dividendRate"]
                    print(f"Market Price: {a}$" + f"\nSector: {b}" + f"\nDividend Rate: {y}%" + f"\nProfit Margin: {c * 100}%")
                elif ticker == "etf":
                    etf = input("Enter ticker: ")
                    data = yf.Ticker(etf)
                    try:
                        a = data.info["regularMarketPrice"]
                        b = data.info["yield"]
                        print(f"{a}$" + f"\n{b * 100}%")
                    except:
                        a = data.info["open"]
                        b = data.info["dividendRate"]
                        print(f"{a}$" + f"\n{b}%")
                elif ticker == "crypto":
                    crypto = input("Enter ticker: ")
                    data = yf.Ticker(crypto)
                    a = data.info["regularMarketPrice"]
                    print("Market price: " + str(a) + "$")
                else:
                    print("incorrect input")

            def value(): #this has its entry widgets in the inv GUI. checks the price x units you hold of an asset and just gives you the total value
                ticker = var.inv5.get()
                shares = var.inv5a.get()
                ticker_data = yf.Ticker(ticker)
                ticker_price = ticker_data.info["regularMarketPrice"]
                output = var.inv6.set(str(ticker_price * shares) + '$')
                

            b1 = tk.Button(window, text='Total', font=('arial', 15, 'bold'), command=t.inv_total).place(x=430, y=355)
            b2 = tk.Button(window, text="Text", command=txt).place(x=540, y=749)  
            b3 = tk.Button(window, text="Save", command=exit).place(x=597, y=749)
            b4 = tk.Button(window, text='Search Basic Investment Information', command=s_info).place(x=657, y=749)
            b5 = tk.Button(window, text='Total Value:', font=('arial', 15), command=value).place(x=30, y=667)
            window.bind('<Return>', enter)
            window.mainloop


    win = windows()

    #new sheets in excel for the next month
    class NEW_SHEET:
        def __init__(self):
            self.wb = openpyxl.load_workbook(filename)
            self.prof_name = tk.StringVar(screen)
            self.page = self.wb.active

        def sheet_window(self):
            window = tk.Toplevel(screen)
            l1 = tk.Label(window, text='NEW_SHEET')
            l1.pack(side='top')
            e1 = tk.Entry(window, textvar=self.prof_name, bd=5)
            e1.pack(side='left')
            def destroy_win():
                self.save_total()
                window.destroy()
            b1 = tk.Button(window, text='ENTER', command=destroy_win)
            b1.pack(side='right')

        def save_total(self):
            self.sheets = self.wb.sheetnames
            ws = self.wb[self.sheets[-1]]
            ws['P2'] = 'INCOME_EARNED: ' + str(t.income)
            ws['P3'] = 'BILLS:' + str(t.bill_tot)
            ws['P4'] = 'TOTAL_SPENT: ' + str(t.grand_total)
            ws['P5'] = 'INCOME_SAVED: ' + str(t.income_saved) + r'%'
            ws['P6'] = 'TOTAL_INVESTED: ' + str(t.investments) 
            ws['P7'] = 'GAS: ' + str(t.gas_tot)
            ws['P8'] = 'FOOD: ' + str(t.food_tot)
            ws['P9'] = 'PAYBACKS: ' + str(t.paybacks_tot)
            ws['P10'] = 'SUPPLIES: ' + str(t.supplies_tot)
            ws['P11'] = 'LUXURY: ' + str(t.luxury_tot)
            ws['P12'] = 'CREDIT: ' + str(t.credit_tot)
            self.wb.save(filename)
            self.new_sheet()

        def new_sheet(self):
            self.wb.create_sheet(self.prof_name.get())
            self.append_to()
            self.wb.save(filename)

        def append_to(self):
            self.sheets = self.wb.sheetnames
            cats = ['Store', 'Item', 'Price', 'Payment_Method', 'Date_Purchased', 'Bills', 
            'GAS', 'FOOD','PAYBACKS', 'SUPPLIES', 'LUXURY', 'CREDIT', 'INCOME', 'TFSA', 'PERSONAL', 'TOTAL']
            ws = self.wb[self.sheets[-1]]
            ws.append(cats)
            ws['C2'] = 0.001
            ws['F2'] = 0.001
            ws['M2'] = 0.001
            ws['N2'] = 0.001
            ws['O2'] = 0.001

    exc = NEW_SHEET()

    #main GUI
    class Main_Window:
        def __init__(self, master):
            self.master = master
            self.L = tk.Label(master, text='Finances', font=('arial', 35, 'bold')).place(x=10, y=2)
            self.La = tk.Label(master, text='Categories', font=('arial', 20)).place(x=80, y=340)
            self.L1 = tk.Label(master, text='Store:', font=('arial', 15)).place(x=87, y=85)
            self.L2 = tk.Label(master, text='Item:', font=('arial', 15)).place(x=96, y=135)
            self.L3 = tk.Label(master, text='Price:', font=('arial', 15)).place(x=90, y=185)
            self.L4 = tk.Label(master, text='P_Method:', font=('arial', 15)).place(x=45, y=235)
            self.L5 = tk.Label(master, text='D_Purchased:', font=('arial', 15)).place(x=17, y=285)
            self.L6 = tk.Label(master, text='Bills:', font=('arial', 17)).place(x=485, y=85)
            self.L7 = tk.Label(master, text="Income_Earned:", font=('arial', 16)).place(x=378, y=250)
            self.L8 = tk.Label(master, text="Necessity's:", font=('arial', 16)).place(x=417, y=300)
            self.L9 = tk.Label(master, text="Discretionnary: ", font=('arial', 16)).place(x=390, y=350)
            self.L10 = tk.Label(master, text='Bills_Payed:', font=('arial', 16)).place(x=416, y=400)
            self.L11 = tk.Label(master, text=r'%_of_Income_Saved:', font=('arial', 16)).place(x=324, y=450)
            self.temp_L1 = tk.Label(master, text='Income:', font=('arial', 17)).place(x=450, y=135)
            self.temp_L2 = tk.Label(master, text='TFSA:', font=('arial', 17)).place(x=465, y=185)
            self.E1 = tk.Entry(master, textvar=var.e_1, bd=5)
            self.E1.place(x=155, y=86)
            self.E2 = tk.Entry(master, textvar=var.e_2, bd=5)
            self.E2.place(x=155, y=136)
            self.E3 = tk.Entry(master, textvar=var.e_3, bd=5)
            self.E3.place(x=155, y=186)
            self.E4 = tk.Entry(master, textvar=var.e_4, bd=5)
            self.E4.place(x=155, y=236)
            self.E5 = tk.Entry(master, textvar=var.e_5, bd=5)
            self.E5.place(x=155, y=286)
            self.E6 = tk.Entry(master, textvar=var.e_6, bd=5)
            self.E6.place(x=540, y=87)
            self.E7 = tk.Entry(master, textvar=var.e_11, bd=5)
            self.E7.place(x=540, y=251)
            self.E8 = tk.Entry(master, textvar=var.e_8, bd=5)
            self.E8.place(x=540, y=301)
            self.E9 = tk.Entry(master, textvar=var.e_9, bd=5)
            self.E9.place(x=540, y=351)
            self.E10 = tk.Entry(master, textvar=var.e_7, bd=5)
            self.E10.place(x=540, y=401)
            self.E11 = tk.Entry(master, textvar=var.e_10, bd=5)
            self.E11.place(x=540, y=451)
            self.E12 = tk.Entry(master, textvar=var.e_13, bd=5)
            self.E12.place(x=540, y=137)
            self.E13 = tk.Entry(master, textvar=var.inv1, bd=5)
            self.E13.place(x=540, y=187)
            self.M1 = tk.Menu(master)
            self.viewM1 = tk.Menu(self.M1, tearoff=0)
            self.viewM1.add_command(label='VIEW_EXCEL', command=file_path)
            self.viewM1.add_command(label='OVERVIEW', command=win.overview_window)
            self.viewM1.add_command(label='INVESTMENTS', command=win.investment_window)
            self.M1.add_cascade(label='VIEW', menu=self.viewM1)
            self.editM1 = tk.Menu(self.M1, tearoff=0)
            self.editM1.add_command(label='NEW_SHEET', command=exc.sheet_window)
            self.editM1.add_command(label='UPLOAD_EXCEL', command=upload_xlsx)
            self.M1.add_cascade(label='EDIT', menu=self.editM1)
            master.bind('<Return>', self.clear)
        
        def clear(self, _event=None):
            info.server_c()
            self.E1.delete(0, 'end')
            self.E2.delete(0, 'end')
            self.E3.delete(0, 'end')
            self.E3.insert(0, 0.0)
            self.E4.delete(0, 'end')
            self.E5.delete(0, 'end')
            self.E6.delete(0, 'end')
            self.E6.insert(0, 0.0)
            self.E8.delete(0, 'end')
            self.E9.delete(0, 'end')
            self.E12.delete(0, 'end')
            self.E12.insert(0 , 0.0)
            self.E13.delete(0, 'end')
            self.E13.insert(0 , 0.0)



    m = Main_Window(screen)

    #class that enters all the information in the excel file, and saves and uploads the files to the server.
    class Enter_Info:
        def __init__(self):
            self.wb = openpyxl.load_workbook(filename)
            self.page = self.wb.active
            self.B1 = tk.Button(screen, text='Total', font=('arial', 13), command=t.main_total).place(x=590, y=2)
            self.B2 = tk.Button(screen, text='Exit', font=('arial', 13), command=self.exit).place(x=660, y=2)
            self.C1 = tk.Checkbutton(screen, text='GAS', variable=var.check1, font=('arial', 15)).place(x=6, y=385)
            self.C2 = tk.Checkbutton(screen, text='FOOD', variable=var.check2, font=('arial', 15)).place(x=6, y=410)
            self.C3 = tk.Checkbutton(screen, text='PAYBACKS', variable=var.check3, font=('arial', 15)).place(x=6, y=435)
            self.C4 = tk.Checkbutton(screen, text='SUPPLIES', variable=var.check4, font=('arial', 15)).place(x=145, y=385)
            self.C5 = tk.Checkbutton(screen, text='LUXURY', variable=var.check5, font=('arial', 15)).place(x=145, y=410)
            self.C6 = tk.Checkbutton(screen, text='WITH_CREDIT', variable=var.check6, font=('arial', 15)).place(x=145, y=435)

        def m_enter(self):
            inf = [var.e_1.get(), var.e_2.get(), var.e_3.get(), var.e_4.get(), var.e_5.get(), var.e_6.get(),
            var.check1.get(), var.check2.get(), var.check3.get(), var.check4.get(), var.check5.get(), var.check6.get(), 
            var.e_13.get(), var.inv1.get(), var.inv2.get()]
            sheets = self.wb.sheetnames
            ws = self.wb[sheets[-1]]
            ws.append(inf)
            self.wb.save(filename)

        def server_c(self):
            dx = Process(target=download_xlsx, name='Server XLSX')
            try:
                dx.start()
                dx.join(timeout=1)
                dx.terminate()
                self.m_enter()
            except:
                self.m_enter()

        def exit(self):
            ux = Process(target=upload_xlsx, name='Server XLSX')
            try:
                ux.start()
                ux.join(timeout=1)
                ux.terminate()
                sys.exit()
            except:
                sys.exit()
            
            



    info = Enter_Info()


    screen.config(menu=m.M1)
    screen.mainloop()

main()
